import torch
import time
import numpy as np
import os
import torch.multiprocessing as mp
from ..env.wrapper import EnvWrapper
import openpyxl

total_steps = 100

class Evaluator:
    def __init__(self, eval_env, eval_per_step: int = 1e4, eval_times: int = 8, cwd: str = '.', print_head = True):
        self.cwd = cwd
        self.env_eval = eval_env
        self.eval_step = 0
        self.total_step = 0
        self.start_time = time.time()
        self.eval_times = eval_times  # number of times that get episodic cumulative return
        self.eval_per_step = eval_per_step  # evaluate the agent per training steps

        self.agent = None
        self.log_max_r = -1e6
        
        self.recorder = []
        if print_head:
            print(f"\n| `step`: Number of samples, or total training steps, or running times of `env.step()`."
                f"\n| `time`: Time spent from the start of training to this moment."
                f"\n| `avgR`: Average value of cumulative rewards, which is the sum of rewards in an episode."
                f"\n| `stdR`: Standard dev of cumulative rewards, which is the sum of rewards in an episode."
                f"\n| `avgS`: Average of steps in an episode."
                f"\n| `objC`: Objective of Critic network. Or call it loss function of critic network."
                f"\n| `objA`: Objective of Actor network. It is the average Q value of the critic network."
                f"\n| {'step':>8}  {'time':>8}  | {'avgR':>8}  {'stdR':>6}  {'avgS':>6}  | {'objC':>8}  {'objA':>8} | {'drop_num':>6} | {'SW_mean':>8} {'SW_std':>8}")

    def evaluate_and_save(self, logging_tuple: tuple, save_net=False):
        # print("开始测试")
        actor = self.agent.act
        actor.eval()
        
        rewards_steps_ary = [get_rewards_and_steps(self.env_eval, actor) for _ in range(self.eval_times)]
        rewards_steps_ary = np.array(rewards_steps_ary)
        info = rewards_steps_ary[:, 2]
        rewards_steps_ary = np.array(rewards_steps_ary[:, :2], dtype=np.float32)
        avg_r = rewards_steps_ary[:, 0].mean() / total_steps  # average of cumulative rewards
        std_r = rewards_steps_ary[:, 0].std() / total_steps  # std of cumulative rewards
        avg_s = rewards_steps_ary[:, 1].mean()  # average of steps in an episode
        avg_drop_num = np.mean([info[i]['drop_num'] for i in range(len(info))])
        avg_sw = np.mean([info[i]['sw'] for i in range(len(info))]) / total_steps
        std_sw = np.std([info[i]['sw'] for i in range(len(info))]) / total_steps

        # print("结束测试")
        
        used_time = time.time() - self.start_time
        self.recorder.append((self.total_step, used_time, avg_r))

        print(f"| {self.total_step:8.2e}  {used_time:8.0f}  "
              f"| {avg_r:8.2f}  {std_r:6.2f}  {avg_s:6.0f}  "
              f"| {logging_tuple[0]:8.2f}  {logging_tuple[1]:8.2f}  "
              f"| {avg_drop_num:4.2f}  "
              f"| {avg_sw:8.2f}  {std_sw:8.2f}")

        if save_net:
            if not os.path.exists("./results"):
                os.makedirs("./results")
            if self.log_max_r <= avg_r:
                self.log_max_r = avg_r
                torch.save(self.agent.act.state_dict(), './results/act_grad.pth')
                torch.save(self.agent.cri.state_dict(), './results/cri_grad.pth')

        file_path = "./results/output.xlsx"

        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
            # 选择默认的活动工作表
            sheet = workbook.active
            # 获取已存在的行数（用于确定增量写入的行号）
            existing_rows = sheet.max_row
        else:
            workbook = openpyxl.Workbook()
            # 选择默认的活动工作表
            sheet = workbook.active
            # 写入标题行
            sheet.append(["Total Step", "Used Time", "Avg R", "Std R", "Avg S", "objC", "objA", "Drop Num", "SW Mean", "SW Std"])

        # 写入数据行
        if os.path.exists(file_path):
            row_data = [self.total_step, used_time, avg_r, std_r, avg_s, logging_tuple[0], logging_tuple[1], avg_drop_num, avg_sw, std_sw]
            # 在已存在的行数上进行增量写入（逐行写入）
            for i, data in enumerate(row_data, start=1):
                sheet.cell(row=existing_rows + 1, column=i, value=data)
        else:
            sheet.append([self.total_step, used_time, avg_r, std_r, avg_s, logging_tuple[0], logging_tuple[1], avg_drop_num, avg_sw, std_sw])

        # 保存工作簿到文件
        workbook.save(file_path)

    
    def test_with_inner_policy(self, policy_id):
        rewards_steps_ary = [step_with_inner_policy(self.env_eval, policy_id) for _ in range(self.eval_times)]
        rewards_steps_ary = np.array(rewards_steps_ary)
        info = rewards_steps_ary[:, 2]
        rewards_steps_ary = np.array(rewards_steps_ary[:, :2], dtype=np.float32)
        avg_r = rewards_steps_ary[:, 0].mean() / total_steps  # average of cumulative rewards
        std_r = rewards_steps_ary[:, 0].std() / total_steps  # std of cumulative rewards
        avg_s = rewards_steps_ary[:, 1].mean()  # average of steps in an episode
        avg_drop_num = np.mean([info[i]['drop_num'] for i in range(len(info))])
        avg_sw = np.mean([info[i]['sw'] for i in range(len(info))]) / total_steps
        std_sw = np.std([info[i]['sw'] for i in range(len(info))]) / total_steps
        
        used_time = time.time() - self.start_time        
        print(f"| {self.total_step}  {used_time:8.0f}  "
              f"| {avg_r:8.2f}  {std_r:6.2f}  {avg_s:6.0f}  "
              f"| None  None  "
              f"| {avg_drop_num:4.2f}  "
              f"| {avg_sw:8.2f}  {std_sw:8.2f}")


def get_rewards_and_steps(env, actor, if_render: bool = False):  # cumulative_rewards and episode_steps
    device = next(actor.parameters()).device  # net.parameters() is a Python generator.

    drop_num = 0
    sw = 0.0

    state = env.reset()
    episode_steps = 0
    cumulative_returns = 0.0  # sum of rewards in an episode
    
    test_PPO_time = False
    if test_PPO_time:
        PPO_time = []   # ms

    for episode_steps in range(total_steps):
        if test_PPO_time: 
            PPO_start_time = time.time()
        tensor_state = torch.as_tensor(state, dtype=torch.float32, device=device).unsqueeze(0)
        tensor_action = actor(tensor_state)
        if test_PPO_time: 
            PPO_end_time = time.time()
            PPO_time.append((PPO_end_time - PPO_start_time)*1000)
            # print(episode_steps, "PPO elapsed time:", (PPO_end_time - PPO_start_time)*1000, device)
            # exit()
        action = tensor_action.detach().cpu().numpy()[0]  # not need detach(), because using torch.no_grad() outside
        state, reward, done, info = env.step(action)
        cumulative_returns += reward

        drop_num += info["drop_num"]
        sw += info["sw"]

        if if_render:
            env.render()
        if done:
            break
    
    if test_PPO_time:
        # 计算平均值和标准差
        mean = np.mean(PPO_time)
        std = np.std(PPO_time)

        # 定义阈值，通常是 Z 分数超过3
        threshold = 3

        # 找到非离群点
        non_outliers = [x for x in PPO_time if (x - mean) < threshold * std]

        # 计算非离群点的平均值
        filtered_mean = np.mean(non_outliers)

        # print("原始数据：", PPO_time)
        print("平均值：", mean)
        # print("去除离群点后的数据：", non_outliers)
        print("平均值（去除离群点后）：", filtered_mean)

    others = {"drop_num":drop_num, "sw":sw}
    return cumulative_returns, episode_steps + 1, others

def step_with_inner_policy(env, policy_id: int):
    env.reset()
    drop_num = 0
    sw = 0.0
    episode_steps = 0
    cumulative_returns = 0.0  # sum of rewards in an episode
    for episode_steps in range(total_steps):
        state, reward, done, info = env.step_with_inner_policy(policy_id)
        cumulative_returns += reward

        drop_num += info["drop_num"]
        sw += info["sw"]

        if done:
            break

    others = {"drop_num":drop_num, "sw":sw}
    return cumulative_returns, episode_steps + 1, others

def test(config):
    config["penalty"] = 0.
    num = 6
    
    evaluators = [Evaluator(eval_env=EnvWrapper(config), eval_times=config["eval_times"], print_head=False) for _ in range(num-1)]
    evaluators.append(Evaluator(eval_env=EnvWrapper(config), eval_times=config["eval_times"]))   # 独立一个出来打印
    
    pool = mp.Pool(processes=10)
    for i in range(num):
        evaluators[i].total_step = i
        pool.apply_async(evaluators[i].test_with_inner_policy, args=(i,))

    pool.close()
    pool.join()