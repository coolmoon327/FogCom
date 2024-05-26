import matplotlib.pyplot as plt
import os
import openpyxl
from statistics import mean
import numpy as np

class ResultCurve:
    def __init__(self):
        self.random_action = -926.99   # 0
        self.random_select = 198.35    # 1
        self.greedy = 1014.68    # 2
        self.all = 470.19  # 4
        self.optimal = 2659.98   # 3
        self.ppo_results = []
        self.time_points = []   # TODO: 这里不太对，应该和 step 统一
        self.steps = 0

        self.draw_var = False

    def set_results(self, results: list):
        self.ppo_results = results
        self.steps = len(self.ppo_results)
        self.time_points = range(0, self.steps)

    def set_points(self, points):
        self.time_points = [int(x) for x in points]

    def update_results(self, new_results: list):
        self.ppo_results += new_results
        self.steps = len(self.ppo_results)
        self.time_points = range(0, self.steps)

    def set_variance(self, ceil_list: list, floor_list: list):
        self.ceil_list = ceil_list
        self.floor_list = floor_list
        self.draw_var = True

    def save_plot(self, filename):
        # 设置中文字体
        plt.rcParams['font.sans-serif'] = ['SimSong']  # 设置中文显示
        plt.rcParams['axes.unicode_minus'] = False    # 解决负号显示问题

        # 设置图形尺寸和标题
        plt.figure(figsize=(12, 8))
        # plt.title("Results Comparison")

        # 绘制基准线和 PPO 奖励曲线
        plt.plot(self.time_points, self.ppo_results, color='blue', linewidth=2, label="PPO 筛选")
        plt.plot(self.time_points, [self.all for _ in range(self.steps)], color='orange', linewidth=3, label="无筛选")
        plt.plot(self.time_points, [self.random_select for _ in range(self.steps)], color='pink', linewidth=3, label="随机选择")
        # plt.plot(self.time_points, [self.optimal for _ in range(self.steps)], color='red', linewidth=3, label="Optimal")
        plt.plot(self.time_points, [self.greedy for _ in range(self.steps)], color='green', linewidth=3, label="贪心选择")
        # plt.plot(self.time_points, [self.random_action for _ in range(self.steps)], label="Random Action")

        if self.draw_var:
            # 填充两根线之间的区域
            plt.fill_between(self.time_points, self.ceil_list, self.floor_list, color='lightblue', alpha=0.5, label="Std")

        # plt.ylim(0, 1800)
        plt.xlim(1, max(self.time_points))
        # plt.xlim(0, 1000)

        # 添加图例和坐标轴标签
        plt.legend(loc='lower right', fontsize=25)
        plt.xlabel("训练次数（20,000 组探索/次训练）", fontsize=35, weight='bold')
        plt.ylabel("单个时隙的平均社会福利值", fontsize=35, weight='bold')

        plt.tick_params(axis='x', labelsize=20)
        plt.tick_params(axis='y', labelsize=20)

        # 保存图像到指定位置
        plt.savefig(filename)
        plt.close()

def draw_PPO(file_path, window_size = 0, print_smoothed=False): # 滑动窗口大小
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        points_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            points_value = row[0]
            # points_data.append(points_value / 10000)    # threads_num * horizon_len
            points_data.append(points_value / 1000) 
        sw_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sw_value = row[8]  # 读取 SW 数据（SW 数据在第 9 列）, 第 9 列的索引为 8
            sw_data.append(sw_value)

        curve = ResultCurve()

        if len(points_data) < window_size or not window_size:
            curve.set_results(sw_data)
            curve.set_points(points_data)
        else:
            # 计算滑动平均值
            smoothed_sw_data = []
            smoothed_ceil_data = []
            smoothed_floor_data = []
            stds = []
            for i in range(len(sw_data) - window_size + 1):
            # for i in range(1295):
                window = sw_data[i : i + window_size]
                smoothed_value = mean(window)
                smoothed_sw_data.append(smoothed_value)

                std = np.std(window)
                stds.append(std)
                ceil_value = smoothed_value + std
                floor_value = smoothed_value - std

                # sorted_window = sorted(window)
                # ceil_value = sorted_window[-2]
                # floor_value = sorted_window[1]

                smoothed_ceil_data.append(ceil_value)
                smoothed_floor_data.append(floor_value)
            
            curve.set_results(smoothed_sw_data)  # 使用滑动平均值数据
            curve.set_points(points_data[:len(smoothed_sw_data)])
            # curve.set_variance(smoothed_ceil_data, smoothed_floor_data)

            if print_smoothed:
                print(smoothed_sw_data[-10:-1], stds[-10:-1])

        dir_name, _ = os.path.split(file_path)
        new_file_path = os.path.join(dir_name, "reward_curve.png")
        curve.save_plot(new_file_path)

if __name__ == "__main__":
    file_path = "../../results/output.xlsx"
    draw_PPO(file_path, 120, True)